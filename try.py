import os
import logging
from flask import Flask, render_template, Response, request, redirect, url_for, flash, session, send_file
from flask_sqlalchemy import SQLAlchemy
from werkzeug.utils import secure_filename
import cv2
import numpy as np
import requests
from datetime import datetime, timedelta
import pandas as pd
import face_recognition
from mtcnn import MTCNN
import torch
from facenet_pytorch import InceptionResnetV1
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# Initialize Flask application
app = Flask(__name__)

# Configuration
app.config['SQLALCHEMY_DATABASE_URI'] = os.getenv('DATABASE_URL', 'sqlite:///employees.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = os.getenv('UPLOAD_FOLDER', 'uploads')
app.secret_key = os.getenv('FLASK_SECRET_KEY', 'your_unique_secret_key_here')

db = SQLAlchemy(app)

# Logging configuration
logging.basicConfig(level=logging.DEBUG)


# Models
class Employee(db.Model):
    __tablename__ = 'employee'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100))
    position = db.Column(db.String(50))
    phone = db.Column(db.String(20))
    email = db.Column(db.String(100))
    job_type = db.Column(db.String(20))
    checkin_time = db.Column(db.String(20))
    checkout_time = db.Column(db.String(20))
    photo_filename = db.Column(db.String(100))
    face_encoding = db.Column(db.LargeBinary)


class Attendance(db.Model):
    __tablename__ = 'attendance'
    id = db.Column(db.Integer, primary_key=True)
    employee_id = db.Column(db.Integer, db.ForeignKey('employee.id'), nullable=False)
    date = db.Column(db.Date, nullable=False)
    status = db.Column(db.String(10), nullable=False)
    checkin_time = db.Column(db.String(10), nullable=True)
    checkout_time = db.Column(db.String(10), nullable=True)


# Initialize face recognition model
facenet_model = InceptionResnetV1(pretrained='vggface2').eval()


def preprocess_image(image):
    image = cv2.resize(image, (160, 160))
    image = image.astype(np.float32) / 255.0
    return torch.tensor(image).unsqueeze(0)


def predict_face(image):
    image = preprocess_image(image)
    embedding = facenet_model(image).detach().numpy()[0]
    return embedding


def compare_embeddings(embedding1, embedding2):
    dot_product = np.dot(embedding1, embedding2)
    norm1 = np.linalg.norm(embedding1)
    norm2 = np.linalg.norm(embedding2)
    return dot_product / (norm1 * norm2)


def load_employee_data():
    employees = Employee.query.all()
    return [(emp.id, emp.name, os.path.join(app.config['UPLOAD_FOLDER'], emp.photo_filename)) for emp in employees]


def save_employee_data(name, position, phone, email, job_type, photo, checkin_time, checkout_time):
    photo_filename = secure_filename(photo.filename) if photo else None
    face_encoding = None
    if photo:
        photo_path = os.path.join(app.config['UPLOAD_FOLDER'], photo_filename)
        photo.save(photo_path)

        # Generate face encoding
        image = face_recognition.load_image_file(photo_path)
        encodings = face_recognition.face_encodings(image)
        if encodings:
            face_encoding = encodings[0].tobytes()

    new_employee = Employee(
        name=name,
        position=position,
        phone=phone,
        email=email,
        job_type=job_type,
        checkin_time=checkin_time,
        checkout_time=checkout_time,
        photo_filename=photo_filename,
        face_encoding=face_encoding
    )
    db.session.add(new_employee)
    db.session.commit()


def record_attendance(employee_id, status, checkin_time=None, checkout_time=None):
    employee = Employee.query.get(employee_id)
    if not employee:
        return

    today = datetime.today().date()
    attendance = Attendance.query.filter_by(employee_id=employee_id, date=today).first()

    if attendance:
        attendance.status = status
        if checkin_time:
            attendance.checkin_time = checkin_time
        if checkout_time:
            attendance.checkout_time = checkout_time
    else:
        attendance = Attendance(
            employee_id=employee_id,
            date=today,
            status=status,
            checkin_time=checkin_time,
            checkout_time=checkout_time
        )
        db.session.add(attendance)
    db.session.commit()


def generate_attendance_report(start_date, end_date):
    attendances = Attendance.query.filter(Attendance.date.between(start_date, end_date)).all()
    employees = Employee.query.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    # Add headers
    headers = ["Employee Name", "Date", "Status", "Check-in Time", "Check-out Time"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    # Add data
    row = 2
    for employee in employees:
        for date in pd.date_range(start_date, end_date):
            attendance = next((a for a in attendances if a.employee_id == employee.id and a.date == date.date()), None)

            status = attendance.status if attendance else "Absent"
            checkin_time = attendance.checkin_time if attendance else "-"
            checkout_time = attendance.checkout_time if attendance else "-"

            ws.append([
                employee.name,
                date.strftime("%Y-%m-%d"),
                status,
                checkin_time,
                checkout_time
            ])

            if status == "Present":
                ws.cell(row=row, column=3).fill = PatternFill(start_color="90EE90", end_color="90EE90",
                                                              fill_type="solid")
            else:
                ws.cell(row=row, column=3).fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1",
                                                              fill_type="solid")

            row += 1

    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    filename = f"attendance_report_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
    wb.save(filename)
    return filename


def get_face_embedding(image):
    face_locations = face_recognition.face_locations(image)
    encodings = face_recognition.face_encodings(image, face_locations)
    return encodings, face_locations


def recognize_face(face_encoding, known_face_encodings, known_face_names, threshold=0.6):
    matches = face_recognition.compare_faces(known_face_encodings, face_encoding, tolerance=threshold)
    face_distances = face_recognition.face_distance(known_face_encodings, face_encoding)

    if len(face_distances) > 0:
        best_match_index = np.argmin(face_distances)
        if matches[best_match_index]:
            employee_id = known_face_names[best_match_index]
            record_attendance(employee_id, "Present", datetime.now().strftime("%H:%M:%S"))
            return [(Employee.query.get(employee_id).name, 1 - face_distances[best_match_index])]

    return [("Unknown", None)]


def load_face_encodings_from_db():
    employees = Employee.query.all()
    known_face_encodings = []
    known_face_names = []
    for employee in employees:
        if employee.face_encoding:
            encoding = np.frombuffer(employee.face_encoding, dtype=np.float64)
            known_face_encodings.append(encoding)
            known_face_names.append(employee.id)
    return known_face_encodings, known_face_names


def generate_frames(camera_url):
    app.logger.info(f"Starting to generate frames for {camera_url}")

    with app.app_context():
        known_face_encodings, known_face_names = load_face_encodings_from_db()

    while True:
        try:
            response = requests.get(camera_url, stream=True)
            if response.status_code != 200:
                app.logger.error(f"Failed to access camera at {camera_url}: {response.status_code}")
                yield (b'--frame\r\n'
                       b'Content-Type: image/jpeg\r\n\r\n' + b'No camera found\r\n')
                break

            bytes_data = bytes()
            for chunk in response.iter_content(chunk_size=1024):
                bytes_data += chunk
                a = bytes_data.find(b'\xff\xd8')
                b = bytes_data.find(b'\xff\xd9')
                if a != -1 and b != -1:
                    jpg = bytes_data[a:b + 2]
                    bytes_data = bytes_data[b + 2:]
                    frame = cv2.imdecode(np.frombuffer(jpg, dtype=np.uint8), cv2.IMREAD_COLOR)

                    if frame is None:
                        app.logger.warning(f"Failed to decode frame from {camera_url}")
                        continue

                    rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                    face_encodings, face_locations = get_face_embedding(rgb_frame)
                    for i, (face_encoding, face_location) in enumerate(zip(face_encodings, face_locations)):
                        x1, y1, x2, y2 = face_location
                        cv2.rectangle(frame, (x1, y1), (x2, y2), (0, 255, 0), 2)
                        recognized = recognize_face(face_encoding, known_face_encodings, known_face_names,
                                                    threshold=0.5)
                        for name, similarity in recognized:
                            label = f"{name}: {similarity:.2f}" if similarity else name
                            cv2.putText(frame, label, (x1, y1 - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 255, 0), 2)

                    _, buffer = cv2.imencode('.jpg', frame)
                    frame = buffer.tobytes()
                    yield (b'--frame\r\n'
                           b'Content-Type: image/jpeg\r\n\r\n' + frame + b'\r\n')

        except Exception as e:
            app.logger.error(f"Error generating frames for {camera_url}: {e}")
            yield (b'--frame\r\n'
                   b'Content-Type: image/jpeg\r\n\r\n' + b'No camera found\r\n')
            break


# Routes
@app.route('/generate_report', methods=['GET', 'POST'])
def generate_report():
    if request.method == 'POST':
        start_date = datetime.strptime(request.form['start_date'], '%Y-%m-%d').date()
        end_date = datetime.strptime(request.form['end_date'], '%Y-%m-%d').date()

        filename = generate_attendance_report(start_date, end_date)
        return send_file(filename, as_attachment=True)

    return render_template('generate_report.html')


@app.route('/update_cameras', methods=['POST'])
def update_cameras():
    num_cameras = int(request.form.get('num_cameras', 0))
    camera_urls = [request.form.get(f'camera_url_{i}', '') for i in range(num_cameras)]
    session['camera_urls'] = camera_urls
    return redirect(url_for('live_stream'))


@app.route('/submit_employee', methods=['POST'])
def submit_employee():
    name = request.form['name']
    position = request.form['position']
    phone = request.form['phone']
    email = request.form['email']
    job_type = request.form['job_type']
    checkin_time = request.form.get('checkin_time') if job_type == 'part_time' else '10:00'
    checkout_time = request.form.get('checkout_time') if job_type == 'part_time' else '17:00'
    photo = request.files['photo']
    try:
        save_employee_data(name, position, phone, email, job_type, photo, checkin_time, checkout_time)
        flash('Employee added successfully!', 'success')
    except Exception as e:
        flash(f'Error adding employee: {e}', 'error')
    return redirect(url_for('index'))


@app.route('/video_feed/<int:camera_id>')
def video_feed(camera_id):
    camera_urls = session.get('camera_urls', [])
    if camera_id < 0 or camera_id >= len(camera_urls):
        return "Camera URL not found", 404

    camera_url = camera_urls[camera_id]

    if not camera_url:
        return "Camera URL not found", 404

    return Response(generate_frames(camera_url),
                    mimetype='multipart/x-mixed-replace; boundary=frame')


@app.route('/live_stream', methods=['GET', 'POST'])
def live_stream():
    if request.method == 'POST':
        if 'num_cameras' in request.form:
            num_cameras = int(request.form.get('num_cameras', 1))
            session['num_cameras'] = num_cameras
            return render_template('index.html', num_cameras=num_cameras, camera_urls=[None] * num_cameras)
        elif 'camera_urls' in request.form:
            num_cameras = int(request.form.get('num_cameras', 1))
            camera_urls = [request.form.get(f'camera_url_{i}', '') for i in range(num_cameras)]
            session['camera_urls'] = camera_urls
            return render_template('index.html', num_cameras=num_cameras, camera_urls=camera_urls)
    num_cameras = session.get('num_cameras', 1)
    camera_urls = session.get('camera_urls', [None] * num_cameras)
    return render_template('index.html', num_cameras=num_cameras, camera_urls=camera_urls)

# ... (previous code remains the same)

@app.route('/check_attendance')
def check_attendance():
    attendances = Attendance.query.all()
    data = [{
        'employee_id': a.employee_id,
        'date': a.date,
        'checkin_time': a.checkin_time,
        'checkout_time': a.checkout_time,
        'status': a.status
    } for a in attendances]

    return {'attendances': data}

@app.route('/favicon.ico')
def favicon():
    return '', 204

@app.route('/add_employee')
def add_employee():
    return render_template('add_employee.html')

@app.route('/report')
def report():
    return render_template('generate_report.html')

@app.route('/')
def index():
    return redirect(url_for('live_stream'))

if __name__ == "__main__":
    with app.app_context():
        db.create_all()
    app.run(host='0.0.0.0', port=5000, debug=True)